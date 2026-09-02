#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""【EVPT】エレベーター特約店管理表 → 銀行提出用 Excel（A3 印刷用）変換スクリプト

Google スプレッドシート「【EVPT】エレベーター特約店管理表」の
「【成約済】EV案件 管理表」シートを Excel でダウンロードし、このスクリプトに渡すと

  * 銀行提出に不要な社内管理列を削除
  * 空欄の行に当たったらそこで打ち切り、「開始月」が「凛」の行は除外
  * 指定したページ数（既定 5 ページ）に収まる範囲で文字・行・列を最大限に拡大
  * A3 横・横 1 ページ幅に収まる印刷設定
  * 見出し行を全ページで繰り返し

した提出用ブックを出力します。元のスプレッドシートは一切変更しません。

使い方:
    python scripts/make_ev_bank_report.py <ダウンロードした.xlsx> [-o 出力フォルダ]
                                          [--date YYMMDD] [--pages 5] [--lines 2]

出力ファイル名:
    【EVPT】エレベーター特約店管理表YYMMDD.xlsx   (YYMMDD = 作成日)
"""

from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import sys
import unicodedata

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

# ---------------------------------------------------------------- 設定
SRC_SHEET = "【成約済】EV案件 管理表"      # 変換元シート名
OUT_SHEET = "EV保守成約案件管理"           # 出力シート名
HEADER_ROW = 4                             # 変換元の見出し行
DOC_TITLE = "EV保守成約案件管理"

# 既定の保存先（Windows の共有ドライブ）。存在しない場合はカレントに保存します。
DEFAULT_OUT_DIR = r"G:\共有ドライブ\★Kevin\☆重要\f\EV関連"
FILE_STEM = "【EVPT】エレベーター特約店管理表"   # 語尾に YYMMDD を付けて保存

TARGET_PAGES = 5          # A3 で何ページに収めるか（--pages で変更）
FONT_NAME = "Meiryo UI"   # 日本語が読みやすく Windows に標準搭載
# 列幅は「COLUMNS の指定 × 幅係数」で決める。係数を下げるほど列が細くなり、
# 紙幅いっぱいに使うぶん文字が大きくなる代わりに折り返しが増えてページ数も増える。
# 目標ページ数に収まる範囲で、いちばん文字が大きくなる係数を自動で選ぶ。
WIDTH_FACTOR_MIN = 0.55
WIDTH_FACTOR_MAX = 1.60
WIDTH_FACTOR_STEP = 0.02
FONT_MIN = 6.0
FONT_MAX = 24.0
MAX_LINES = 2             # 1 セルの折り返し上限（--lines で変更）
# 列幅の何倍を超えたら折り返すか。これ以下の超過は 1 行のまま自動縮小で見せる
# （1.4 なら、縮小率が 7 割を下回るときだけ 2 行にする）
WRAP_THRESHOLD = 1.4
LINE_RATIO = 1.32         # 1 行の高さ ÷ 文字サイズ
ROW_PAD_PT = 1.5          # 行の上下余白

PAPER_A3 = 8
ORIENTATION = "landscape"     # A3 横。縦にしたい場合は "portrait"
MARGIN_SIDE_IN = 0.25
MARGIN_TOPBOTTOM_IN = 0.25

# 銀行提出用に残す列。(出力見出し, 変換元見出し, 幅[半角文字数], 表示形式, 折返し, 寄せ)
# 変換元見出しで列を探すため、元シートで列が増減・移動しても追従します。
# 幅は文字サイズに比例して決まるので、ここは「半角何文字ぶん見せたいか」で指定します。
COLUMNS = [
    ("成約Code",     "成約Code",       8, None,      False, "center"),
    ("担当",         "担当",           8, None,      False, "center"),
    ("顧客名",       "顧客名",        32, None,      True,  "left"),
    ("物件名",       "物件名",        32, None,      True,  "left"),
    ("開始月",       "開始月",         8, "yyyy/mm", False, "center"),
    ("弊社\n入金月", "弊社\n入金月",   8, "yyyy/mm", False, "center"),
    ("提案内容",     "提案内容",       9, None,      False, "center"),
    ("基数",         "基数",           6, '0"基"',   False, "center"),
    ("契約形態",     "契約形態",       7, None,      False, "center"),
    ("点検頻度",     "点検頻度",       8, None,      False, "center"),
    ("OP",           "OP",             7, None,      False, "center"),
    ("仕切",         "仕切",          12, "¥#,##0",  False, "right"),
    ("見積料金",     "見積料金",      12, "¥#,##0",  False, "right"),
    ("❶EVPT\n売上", "❶EVPT\n売上",  10, "¥#,##0",  False, "right"),
    ("❷PTI",        "❷PTI",         10, "¥#,##0",  False, "right"),
    ("❸EVPT NP",    "❸EVPT NP",     10, "¥#,##0",  False, "right"),
    ("結果",         "結果",          11, None,      False, "center"),
    ("備考",         "備考",          27, None,      True,  "left"),
]
# 変換元にあっても銀行提出では落とす列（社内管理用）:
#   PT① / 担当(2つ目) / 既存契約 / 仕様 / ベンダー / 既存料金 / 結果番号 /
#   Z列以降の集計ダッシュボード

# 除外条件: 「開始月」がこの値の行は出力しない
EXCLUDE_START_MONTH = {"凛"}

ERROR_TEXTS = {"#VALUE!", "#N/A", "#REF!", "#NAME?", "#DIV/0!", "#NULL!", "#NUM!"}

# 列のまとまり。見出しの色を分け、境目に太めの縦罫線を引く
#   (グループ名, 先頭列の出力見出し, 見出しの地色)
COLUMN_GROUPS = [
    ("案件",   "成約Code", "1F3864"),
    ("時期",   "開始月",   "2E5A8A"),
    ("契約",   "提案内容", "2F6E6A"),
    ("金額",   "仕切",     "1E5F3F"),
    ("結果",   "結果",     "5A5F6E"),
]

# 担当ごとの色。ここに無い担当は OWNER_PALETTE から自動で割り当てます
OWNER_FILLS = {
    "小坂":           "DCE9F7",   # 青
    "山内":           "E3F1DE",   # 緑
    "河内山":         "FCE7D6",   # 橙
    "ライクタイガー": "EAE1F5",   # 紫
    "啓就":           "FBF0C9",   # 黄
    "諸富":           "DDEFF0",   # 水
    "上野":           "F9DFE6",   # 桃
}
OWNER_PALETTE = ["E7E3D8", "E2E8F0", "F0E6E0", "E6EEE2", "EFE6EF", "E0EAEA"]

# 「結果」ごとの色分け（地色, 文字色, 太字）
RESULT_STYLES = {
    "保守成約":   ("D9EAD3", "1E5F3F", True),
    "リニュ成約": ("D9EAD3", "1E5F3F", True),
    "新設成約":   ("D9EAD3", "1E5F3F", True),
    "進捗中":     ("FCEFC7", "8A6100", True),
    "保留":       ("FCEFC7", "8A6100", True),
    "失注":       ("EAEAEA", "6B6B6B", False),
    "解約":       ("F7D9D9", "A33A3A", True),
}

HEADER_FONT_COLOR = "FFFFFF"
BAND_FILL = "EFF5FB"        # 1 行おきの薄い地色（横方向に目で追いやすくする）
ZERO_FONT = "A0A0A0"        # ¥0 は控えめに
MUTED_FONT = "8A8A8A"       # 失注・解約の行は文字を落ち着かせる
MUTED_RESULTS = {"失注", "解約"}
GRID = Side(style="thin", color="A8BACD")
GROUP_LINE = Side(style="medium", color="53708C")   # 列グループの境目
YEAR_LINE = Side(style="medium", color="53708C")    # 開始年が変わる区切り
BORDER = Border(left=GRID, right=GRID, top=GRID, bottom=GRID)

# Excel の列幅 1 単位 = 既定フォント(Calibri 11)の数字幅 7px。px = width*7 + 5
PX_PER_UNIT = 7.0
CELL_PAD_PX = 5
DPI = 96


# ---------------------------------------------------------------- ユーティリティ
def page_size_in() -> tuple[float, float]:
    """A3 の印字可能サイズ（インチ）。"""
    w, h = (16.54, 11.69) if ORIENTATION == "landscape" else (11.69, 16.54)
    return w - MARGIN_SIDE_IN * 2, h - MARGIN_TOPBOTTOM_IN * 2


def char_px(font_size: float) -> float:
    """半角 1 文字ぶんの幅（px）。"""
    return font_size * (DPI / 72) * 0.5


def display_len(text) -> int:
    """半角=1 / 全角=2 で数えた表示幅。"""
    return sum(2 if unicodedata.east_asian_width(ch) in "WFA" else 1 for ch in str(text))


def norm(text) -> str:
    """見出し比較用に空白・改行を潰した文字列。"""
    return re.sub(r"\s+", "", str(text or ""))


def clean(value):
    """数式エラー文字列と空文字を None にそろえる。"""
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in ERROR_TEXTS or stripped == "":
            return None
    return value


def as_text(value, number_format: str | None) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return "0000/00"
    if isinstance(value, (int, float)):
        return f"¥{value:,.0f}" if "¥" in (number_format or "") else f"{value:,.0f}"
    return str(value)


def find_columns(ws) -> dict:
    """変換元の見出し行から {正規化見出し: 列番号} を作る（最初に出た列を採用）。"""
    found = {}
    for col in range(1, ws.max_column + 1):
        key = norm(ws.cell(HEADER_ROW, col).value)
        if key and key not in found:
            found[key] = col
    return found


def collect_rows(ws, header_map: dict) -> tuple[list[list], int, int]:
    """出力対象の明細を集める。

    - 見出し行の次から下へ走査し、**出力する 18 列がすべて空の行に当たったら
      そこで打ち切る**（それ以降は印刷対象外）。
    - 「開始月」が EXCLUDE_START_MONTH の行は飛ばす。

    戻り値: (行データ, 打ち切った行番号, 除外した件数)
    """
    src_cols = [header_map[norm(h)] for _l, h, *_ in COLUMNS]
    start_col = header_map[norm("開始月")]
    rows, excluded, stopped_at = [], 0, ws.max_row + 1
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        values = [clean(ws.cell(r, c).value) for c in src_cols]
        if all(v is None for v in values):
            stopped_at = r
            break
        if clean(ws.cell(r, start_col).value) in EXCLUDE_START_MONTH:
            excluded += 1
            continue
        rows.append(values)
    return rows, stopped_at, excluded


def line_count(text: str, capacity_hw: int, wrap: bool) -> int:
    """このセルを何行で表示するか。

    折り返し可の列でも、列幅の WRAP_THRESHOLD 倍までの超過は 1 行のままにして
    自動縮小で見せる（軽い超過で行が倍の高さになるのを防ぐ）。
    それを超える長文だけ MAX_LINES 行まで折り返す。
    """
    if not text or not wrap or MAX_LINES <= 1:
        return 1
    lines = 0
    for part in text.split("\n"):
        length = display_len(part)
        lines += max(1, -(-length // capacity_hw)) if length > capacity_hw * WRAP_THRESHOLD else 1
    return max(1, min(MAX_LINES, lines))


def layout(rows: list[list], width_factor: float) -> dict:
    """ある列幅係数での列幅・文字サイズ・行高・ページ数を見積もる。

    列幅は必ず A3 の紙幅いっぱいになるようにし（縮小率 100%）、
    そこへ収まる文字サイズを逆算する。係数が小さいほど文字は大きくなる。
    """
    page_w_in, page_h_in = page_size_in()
    page_w_px = page_w_in * DPI

    caps = [max(3, round(cap * width_factor)) for _l, _s, cap, *_ in COLUMNS]
    # 紙幅 = Σ(半角文字数 × 半角1文字の幅) + 列ごとの余白
    font_size = (page_w_px - len(caps) * 10) / (sum(caps) * (DPI / 72) * 0.5)
    font_size = max(FONT_MIN, min(FONT_MAX, round(font_size * 2) / 2))

    cpx = char_px(font_size)
    widths_px = [cap * cpx + 10 for cap in caps]
    total_px = sum(widths_px)
    if total_px < page_w_px:
        # 文字数を整数に丸めた端数ぶんを配分し、紙幅をぴったり使い切る
        stretch = page_w_px / total_px
        widths_px = [w * stretch for w in widths_px]
        caps = [max(1, int((w - 10) / cpx)) for w in widths_px]
        total_px = page_w_px
    scale = min(1.0, page_w_px / total_px)

    line_pt = font_size * LINE_RATIO
    title_h = round((font_size + 6) * LINE_RATIO + 6, 1)
    meta_h = round((font_size + 1) * LINE_RATIO + 4, 1)
    head_h = round(2 * line_pt + 6, 1)
    repeat_pt = head_h            # 2 ページ目以降に繰り返すのは見出し行だけ

    row_heights, multiline = [], 0
    for values in rows:
        lines = 1
        for value, cap, (_l, _s, _c, numfmt, wrap, _a) in zip(values, caps, COLUMNS):
            lines = max(lines, line_count(as_text(value, numfmt), cap, wrap))
        multiline += lines > 1
        row_heights.append(round(lines * line_pt + ROW_PAD_PT, 1))

    # 実際の改ページと同じ要領で数える（1 ページ目だけ表題ぶん減る）
    usable_pt = page_h_in * 72 / scale - repeat_pt
    pages, filled = 1, title_h + meta_h
    for h in row_heights:
        if filled + h > usable_pt:
            pages += 1
            filled = 0.0
        filled += h

    return {
        "width_factor": width_factor,
        "font_size": font_size,
        "widths_px": widths_px,
        "total_px": total_px,
        "scale": scale,
        "row_heights": row_heights,
        "title_h": title_h,
        "meta_h": meta_h,
        "head_h": head_h,
        "caps": caps,
        "pages": pages,
        "multiline": multiline,
    }


def fit_layout(rows: list[list], target_pages: int) -> dict:
    """目標ページ数に収まる範囲で、いちばん文字が大きくなる列幅係数を選ぶ。"""
    factor = WIDTH_FACTOR_MIN
    best = layout(rows, factor)
    while best["pages"] > target_pages and factor + WIDTH_FACTOR_STEP <= WIDTH_FACTOR_MAX:
        factor = round(factor + WIDTH_FACTOR_STEP, 2)
        best = layout(rows, factor)
    return best


# ---------------------------------------------------------------- 変換本体
def build(src_path: str, out_path: str, made_on: dt.date, target_pages: int) -> dict:
    src_wb = load_workbook(src_path, data_only=True)
    if SRC_SHEET not in src_wb.sheetnames:
        raise SystemExit(
            f"シート「{SRC_SHEET}」が見つかりません。含まれるシート: {src_wb.sheetnames}"
        )
    src = src_wb[SRC_SHEET]

    header_map = find_columns(src)
    missing = [h for _l, h, *_ in COLUMNS if norm(h) not in header_map]
    if missing:
        raise SystemExit("変換元に見出しが見つかりません: " + " / ".join(missing))

    rows, stopped_at, excluded = collect_rows(src, header_map)
    if not rows:
        raise SystemExit("出力対象の明細が 1 件もありません。")

    lay = fit_layout(rows, target_pages)
    fs = lay["font_size"]
    result_idx = [i for i, (label, *_r) in enumerate(COLUMNS) if label == "結果"][0]

    wb = Workbook()
    ws = wb.active
    ws.title = OUT_SHEET
    ncols = len(COLUMNS)
    last_letter = get_column_letter(ncols)

    # --- 表題（全ページで繰り返す 1〜3 行目）
    ws.merge_cells(f"A1:{last_letter}1")
    ws["A1"] = DOC_TITLE
    ws["A1"].font = Font(name=FONT_NAME, size=fs + 6, bold=True, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = lay["title_h"]

    ws.merge_cells(f"A2:{last_letter}2")
    ws["A2"] = (
        f"作成日：{made_on.strftime('%Y年%m月%d日')}　／　件数：{len(rows)} 件"
        f"　／　出典：【EVPT】エレベーター特約店管理表「{SRC_SHEET}」"
    )
    ws["A2"].font = Font(name=FONT_NAME, size=fs + 1)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = lay["meta_h"]

    # --- 列グループ（見出しの色分けと縦罫線の位置）
    labels = [c[0] for c in COLUMNS]
    group_starts, group_fill = {}, {}
    current = COLUMN_GROUPS[0][2]
    for _name, first_label, color in COLUMN_GROUPS:
        if first_label in labels:
            group_starts[labels.index(first_label) + 1] = True
    for i, label in enumerate(labels, start=1):
        for _name, first_label, color in COLUMN_GROUPS:
            if label == first_label:
                current = color
        group_fill[i] = current

    def edges(col: int, top: Side, bottom: Side) -> Border:
        """列グループの境目だけ縦罫線を太くする。"""
        return Border(
            left=GROUP_LINE if col in group_starts and col > 1 else GRID,
            right=GROUP_LINE if col + 1 in group_starts else GRID,
            top=top, bottom=bottom,
        )

    # --- 見出し行
    hrow = 3
    for i, label in enumerate(labels, start=1):
        cell = ws.cell(hrow, i, label)
        cell.font = Font(name=FONT_NAME, size=fs, bold=True, color=HEADER_FONT_COLOR)
        cell.fill = PatternFill("solid", fgColor=group_fill[i])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = edges(i, GROUP_LINE, GROUP_LINE)
    ws.row_dimensions[hrow].height = lay["head_h"]

    # --- 明細
    owner_idx = labels.index("担当")
    start_idx = labels.index("開始月")
    owner_colors = dict(OWNER_FILLS)
    spare = iter(OWNER_PALETTE)
    for values in rows:                       # 未登録の担当にも色を用意する
        owner = values[owner_idx]
        if owner and owner not in owner_colors:
            owner_colors[owner] = next(spare, "EDEDED")

    def year_key(value):
        return value.year if isinstance(value, (dt.datetime, dt.date)) else str(value or "")

    out_row = hrow
    prev_year = object()
    for n, values in enumerate(rows):
        out_row += 1
        result = values[result_idx]
        # 開始年が変わる行には区切り線を引く（年ごとのまとまりが分かる）
        this_year = year_key(values[start_idx])
        new_block = this_year != prev_year
        if new_block:
            prev_year = this_year
        band = BAND_FILL if n % 2 else None      # 1 行おきの縞
        top = YEAR_LINE if new_block and n else GRID

        for i, (value, (label, _s, _cap, numfmt, wrap, halign)) in enumerate(
            zip(values, COLUMNS), start=1
        ):
            cell = ws.cell(out_row, i, value)
            # 失注・解約の行は文字を薄くして、生きている契約を目立たせる
            color = MUTED_FONT if result in MUTED_RESULTS else None
            bold = label == "❸EVPT NP" and result not in MUTED_RESULTS
            fill = band

            if label == "担当" and value in owner_colors:
                fill = owner_colors[value]
            elif label == "結果" and result in RESULT_STYLES:
                fill, color, bold = RESULT_STYLES[result]
            elif numfmt and "¥" in numfmt and value == 0:
                color = ZERO_FONT                     # ¥0 は目立たせない

            cell.font = Font(name=FONT_NAME, size=fs, bold=bold, color=color)
            # 大きく超過する長文だけ折り返し、それ以外は 1 行のまま自動縮小
            do_wrap = line_count(as_text(value, numfmt), lay["caps"][i - 1], wrap) > 1
            cell.alignment = Alignment(
                horizontal=halign, vertical="center",
                wrap_text=do_wrap, shrink_to_fit=not do_wrap,
            )
            cell.border = edges(i, top, GRID)
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
            if numfmt and isinstance(value, (int, float, dt.datetime, dt.date)):
                cell.number_format = numfmt
        ws.row_dimensions[out_row].height = lay["row_heights"][n]

    # --- 列幅
    for i, width_px in enumerate(lay["widths_px"], start=1):
        ws.column_dimensions[get_column_letter(i)].width = round(
            (width_px - CELL_PAD_PX) / PX_PER_UNIT, 2
        )

    # --- 表示・印刷設定
    ws.freeze_panes = ws.cell(hrow + 1, 5)          # 見出し行と 顧客名/物件名 までを固定
    ws.sheet_view.showGridLines = False

    ws.page_setup.paperSize = PAPER_A3
    ws.page_setup.orientation = ORIENTATION
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0                    # 高さはページ数なりで
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = f"{hrow}:{hrow}"           # 見出し行を毎ページ印刷
    ws.print_area = f"A1:{last_letter}{out_row}"     # 明細の最終行までが印刷対象
    ws.page_margins.left = ws.page_margins.right = MARGIN_SIDE_IN
    ws.page_margins.top = ws.page_margins.bottom = MARGIN_TOPBOTTOM_IN
    ws.page_margins.header = ws.page_margins.footer = 0.15
    ws.oddFooter.left.text = DOC_TITLE + "　" + made_on.strftime("%Y/%m/%d")
    ws.oddFooter.left.size = 9
    ws.oddFooter.right.text = "&P / &N"
    ws.oddFooter.right.size = 9

    wb.save(out_path)

    return {
        "path": out_path,
        "count": len(rows),
        "excluded": excluded,
        "stopped_at": stopped_at,
        "font_size": fs,
        "scale": lay["scale"],
        "pages": lay["pages"],
    }


# ---------------------------------------------------------------- CLI
def resolve_out_dir(requested: str | None) -> str:
    """保存先を決める。既定は共有ドライブで、無ければ作りにいく。"""
    if requested:
        os.makedirs(requested, exist_ok=True)
        return requested

    # Windows 以外で "G:\..." を渡すと、カレントに変な名前のフォルダができてしまう
    windows_path = re.match(r"^[A-Za-z]:[\\/]", DEFAULT_OUT_DIR) is not None
    if not windows_path or os.name == "nt":
        try:
            os.makedirs(DEFAULT_OUT_DIR, exist_ok=True)   # 無ければ作る
            return DEFAULT_OUT_DIR
        except OSError as err:
            print(f"※ 既定の保存先を使えません（{err}）", file=sys.stderr)
    else:
        print("※ Windows ではないため既定の保存先を使えません", file=sys.stderr)

    print(
        f"※ カレントフォルダに保存します。共有ドライブに置く場合は"
        f' -o "{DEFAULT_OUT_DIR}" を付けて実行してください。',
        file=sys.stderr,
    )
    return os.getcwd()


def main() -> None:
    ap = argparse.ArgumentParser(description="EV案件管理表を銀行提出用 A3 Excel に変換します")
    ap.add_argument("src", help="Google スプレッドシートからダウンロードした .xlsx")
    ap.add_argument("-o", "--out-dir", default=None, help=f"保存先フォルダ（既定: {DEFAULT_OUT_DIR}）")
    ap.add_argument("--date", default=None, help="ファイル名の日付 YYMMDD（既定: 本日）")
    ap.add_argument(
        "--lines", type=int, default=MAX_LINES,
        help=f"長い文字列を何行まで折り返すか（既定: {MAX_LINES}。1 なら折り返さず自動縮小）",
    )
    ap.add_argument(
        "--pages", type=int, default=TARGET_PAGES,
        help=f"A3 で何ページに収めるか（既定: {TARGET_PAGES}。増やすほど文字が大きくなります）",
    )
    args = ap.parse_args()

    if args.date:
        if not re.fullmatch(r"\d{6}", args.date):
            raise SystemExit("--date は YYMMDD の 6 桁で指定してください（例 260511）")
        made_on = dt.datetime.strptime(args.date, "%y%m%d").date()
    else:
        made_on = dt.date.today()
    if args.pages < 1:
        raise SystemExit("--pages は 1 以上で指定してください")
    if args.lines < 1:
        raise SystemExit("--lines は 1 以上で指定してください")
    globals()["MAX_LINES"] = args.lines

    out_dir = resolve_out_dir(args.out_dir)
    out_path = os.path.join(out_dir, f"{FILE_STEM}{made_on.strftime('%y%m%d')}.xlsx")

    info = build(args.src, out_path, made_on, args.pages)
    printed = info["font_size"] * info["scale"]
    print(f"作成しました: {info['path']}")
    print(f"  件数        : {info['count']} 件"
          f"（{info['stopped_at']} 行目が空欄のため、それ以降は対象外）")
    if info["excluded"]:
        print(f"  除外        : 開始月が {'/'.join(sorted(EXCLUDE_START_MONTH))} の {info['excluded']} 件")
    print(f"  用紙        : A3 {'横' if ORIENTATION == 'landscape' else '縦'} / 横 1 ページ幅")
    print(f"  ページ数    : {info['pages']} ページ（目標 {args.pages} ページ以内）")
    if info["pages"] > args.pages:
        print(f"  ※ {args.pages} ページには収まりませんでした"
              f"（文字サイズの下限 {FONT_MIN:g}pt に到達）。列を減らすか --pages を増やしてください。",
              file=sys.stderr)
    print(f"  文字        : {info['font_size']:g}pt × 縮小率 {info['scale'] * 100:.0f}%"
          f" → 紙の上で 約 {printed:.1f}pt")


if __name__ == "__main__":
    main()
