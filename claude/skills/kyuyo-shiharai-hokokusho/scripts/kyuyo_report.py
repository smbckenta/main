#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""給与支払報告書（個人別明細書・総括表）を Excel で作る。

    python kyuyo_report.py init  出力先.xlsx [--year 2025]
        入力用ブックを作る（支払者 / 受給者 の 2 シート）。

    python kyuyo_report.py build 入力.xlsx --out 報告書.xlsx
        入力用ブックを読んで、個人別明細書と総括表を作る。

金額の考え方:
  「給与所得控除後の金額」と「所得控除の額の合計額」は、空欄なら参考値を計算して
  入れる。年末調整の結果が手元にあるなら、そちらを入力欄に直接書いたほうが確実。
  計算根拠は reference.md に書いてある。
"""

import argparse
import sys
import unicodedata
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

# ---------------------------------------------------------------- 入力欄の定義

# (列名, 幅, 説明)。説明は入力ブックのコメント行に出す。
PAYER_FIELDS = [
    ("提出先市区町村", "受給者ごとに違うので、ここは代表的な提出先だけ書けばよい（総括表は自動で市区町村ごとに作る）"),
    ("指定番号", "特別徴収義務者指定番号。市区町村ごとに違う場合は受給者シート側で上書きできる"),
    ("個人番号又は法人番号", "13桁の法人番号、または事業主の個人番号"),
    ("所在地郵便番号", ""),
    ("所在地", ""),
    ("名称", "会社名・屋号"),
    ("代表者職氏名", ""),
    ("電話番号", ""),
    ("連絡者の係名", "経理課 など"),
    ("連絡者の氏名", ""),
    ("連絡者の電話番号", ""),
    ("事業種目", ""),
    ("所轄税務署", ""),
    ("給与支払の方法及び期日", "月給・毎月25日 など"),
    ("納入書の送付", "要 / 不要"),
    ("関与税理士等の氏名", ""),
    ("関与税理士等の電話番号", ""),
]

# (列名, 幅, 種別)。種別 num は数値、date は日付文字列、str はそれ以外。
EMPLOYEE_FIELDS = [
    ("受給者番号", 12, "str"),
    ("氏名", 16, "str"),
    ("フリガナ", 16, "str"),
    ("個人番号", 16, "str"),
    ("生年月日", 13, "date"),
    ("役職名", 12, "str"),
    ("提出先市区町村", 16, "str"),
    ("指定番号", 12, "str"),
    ("徴収区分", 22, "str"),
    ("郵便番号", 11, "str"),
    ("住所又は居所（1月1日現在）", 34, "str"),
    ("種別", 12, "str"),
    ("支払金額", 13, "num"),
    ("給与所得控除後の金額", 18, "num"),
    ("所得控除の額の合計額", 18, "num"),
    ("源泉徴収税額", 13, "num"),
    ("控除対象配偶者の有無等", 20, "str"),
    ("配偶者（特別）控除の額", 20, "num"),
    ("配偶者の合計所得", 15, "num"),
    ("扶養親族_特定", 13, "num"),
    ("扶養親族_老人", 13, "num"),
    ("扶養親族_老人のうち同居老親等", 26, "num"),
    ("扶養親族_その他", 15, "num"),
    ("16歳未満扶養親族の数", 20, "num"),
    ("特定親族特別控除の額", 20, "num"),
    ("障害者_特別（本人を除く）", 24, "num"),
    ("障害者_その他（本人を除く）", 24, "num"),
    ("障害者_同居特別（本人を除く）", 26, "num"),
    ("非居住者である親族の数", 20, "num"),
    ("社会保険料等の金額", 17, "num"),
    ("うち小規模企業共済等掛金", 24, "num"),
    ("生命保険料の控除額", 17, "num"),
    ("地震保険料の控除額", 17, "num"),
    ("住宅借入金等特別控除の額", 22, "num"),
    ("基礎控除の額", 13, "num"),
    ("所得金額調整控除額", 17, "num"),
    ("新生命保険料の金額", 17, "num"),
    ("旧生命保険料の金額", 17, "num"),
    ("介護医療保険料の金額", 18, "num"),
    ("新個人年金保険料の金額", 20, "num"),
    ("旧個人年金保険料の金額", 20, "num"),
    ("国民年金保険料等の金額", 20, "num"),
    ("旧長期損害保険料の金額", 20, "num"),
    ("住宅借入金等特別控除適用数", 24, "num"),
    ("居住開始年月日1", 15, "date"),
    ("住宅借入金等特別控除区分1", 24, "str"),
    ("住宅借入金等年末残高1", 20, "num"),
    ("居住開始年月日2", 15, "date"),
    ("住宅借入金等特別控除区分2", 24, "str"),
    ("住宅借入金等年末残高2", 20, "num"),
    ("本人が障害者", 14, "str"),
    ("寡婦", 8, "str"),
    ("ひとり親", 10, "str"),
    ("勤労学生", 10, "str"),
    ("未成年者", 10, "str"),
    ("外国人", 8, "str"),
    ("死亡退職", 10, "str"),
    ("災害者", 8, "str"),
    ("乙欄", 8, "str"),
    ("中途就職年月日", 15, "date"),
    ("中途退職年月日", 15, "date"),
    ("摘要", 40, "str"),
]

# 入力例。init のときに 1 行だけ薄く入れておく。
SAMPLE_EMPLOYEE = {
    "受給者番号": "0001",
    "氏名": "山田 太郎",
    "フリガナ": "ヤマダ タロウ",
    "生年月日": "1990-04-01",
    "提出先市区町村": "福岡市",
    "徴収区分": "特別徴収",
    "郵便番号": "812-0011",
    "住所又は居所（1月1日現在）": "福岡県福岡市博多区博多駅前1-1-1",
    "種別": "給料・賞与",
    "支払金額": 4800000,
    "源泉徴収税額": 98000,
    "社会保険料等の金額": 720000,
    "控除対象配偶者の有無等": "無",
    "扶養親族_その他": 0,
}

TOKUBETSU = "特別徴収"
FUTSU_TAISHOKU = "普通徴収（退職者）"
FUTSU_SONOTA = "普通徴収（退職者を除く）"
COLLECTION_KINDS = [TOKUBETSU, FUTSU_TAISHOKU, FUTSU_SONOTA]

# --------------------------------------------------------------- 見た目の共通部品

THIN = Side(style="thin", color="808080")
MEDIUM = Side(style="medium", color="000000")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LABEL_FILL = PatternFill("solid", fgColor="EFEFEF")
NOTE_FONT = Font(size=8, color="808080")
LABEL_FONT = Font(size=8)
VALUE_FONT = Font(size=10)
TITLE_FONT = Font(size=13, bold=True)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

FORM_COLS = 24  # 個人別明細書 1 枚あたりの列数


def _yen(value):
    """空欄と 0 を区別したいので、None はそのまま返す。"""
    return None if value in (None, "") else value


def cell(ws, row, col, span, value, *, label=False, money=False, align=None):
    """(row, col) から span 列ぶんを結合して 1 マスにする。"""
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
    c = ws.cell(row=row, column=col)
    c.value = value
    c.font = LABEL_FONT if label else VALUE_FONT
    if label:
        c.fill = LABEL_FILL
    if align is not None:
        c.alignment = align
    else:
        c.alignment = CENTER if label else (RIGHT if money else LEFT)
    if money and value not in (None, ""):
        c.number_format = "#,##0"
    for i in range(span):
        ws.cell(row=row, column=col + i).border = BOX
    return c


def row_of(ws, row, cells, *, label=False):
    """[(span, value), ...] を 1 行に並べる。"""
    col = 1
    for span, value in cells:
        cell(ws, row, col, span, value, label=label)
        col += span
    return col


# ------------------------------------------------------- 給与所得控除後の金額の計算


def salary_income(payment, year):
    """給与所得控除後の給与等の金額（所得税法別表第五の考え方）。

    令和 7 年分（2025 年分）から給与所得控除の最低保障額が 55 万円→65 万円に
    引き上げられたので、年で表を切り替える。
    """
    a = int(payment or 0)
    if a <= 0:
        return 0

    if year >= 2025:  # 令和7年分以降
        if a < 1_900_000:
            return max(0, a - 650_000)
        if a < 3_600_000:
            b = (a // 4) // 1000 * 1000
            return int(b * 2.8) - 80_000
        if a < 6_600_000:
            b = (a // 4) // 1000 * 1000
            return int(b * 3.2) - 440_000
        if a < 8_500_000:
            return int(a * 0.9) - 1_100_000
        return a - 1_950_000

    # 令和2年分〜令和6年分
    if a < 551_000:
        return 0
    if a < 1_619_000:
        return a - 550_000
    if a < 1_620_000:
        return 1_069_000
    if a < 1_622_000:
        return 1_070_000
    if a < 1_624_000:
        return 1_072_000
    if a < 1_628_000:
        return 1_074_000
    if a < 1_800_000:
        b = (a // 4) // 1000 * 1000
        return int(b * 2.4) + 100_000
    if a < 3_600_000:
        b = (a // 4) // 1000 * 1000
        return int(b * 2.8) - 80_000
    if a < 6_600_000:
        b = (a // 4) // 1000 * 1000
        return int(b * 3.2) - 440_000
    if a < 8_500_000:
        return int(a * 0.9) - 1_100_000
    return a - 1_950_000


def deduction_total(emp):
    """所得控除の額の合計額の参考値。入力された控除額をそのまま足すだけ。

    扶養控除・障害者控除は人数から金額を出す。基礎控除は入力がなければ足さない
    （年によって額が変わるうえ、本人の合計所得で変わるため）。
    """

    def n(key):
        v = emp.get(key)
        return int(v) if isinstance(v, (int, float)) else 0

    total = 0
    total += n("社会保険料等の金額")
    total += n("生命保険料の控除額")
    total += n("地震保険料の控除額")
    total += n("配偶者（特別）控除の額")
    total += n("特定親族特別控除の額")
    total += n("基礎控除の額")

    # 扶養控除
    total += n("扶養親族_特定") * 630_000
    doukyo_rooshin = n("扶養親族_老人のうち同居老親等")
    total += doukyo_rooshin * 580_000
    total += max(0, n("扶養親族_老人") - doukyo_rooshin) * 480_000
    total += n("扶養親族_その他") * 380_000

    # 障害者控除（扶養親族等の分）
    total += n("障害者_同居特別（本人を除く）") * 750_000
    total += max(0, n("障害者_特別（本人を除く）") - n("障害者_同居特別（本人を除く）")) * 400_000
    total += n("障害者_その他（本人を除く）") * 270_000

    # 本人に係る控除
    honnin = str(emp.get("本人が障害者") or "")
    if "特別" in honnin:
        total += 400_000
    elif honnin.strip() in ("○", "有", "一般", "その他"):
        total += 270_000
    if str(emp.get("寡婦") or "").strip() in ("○", "有"):
        total += 270_000
    if str(emp.get("ひとり親") or "").strip() in ("○", "有"):
        total += 350_000
    if str(emp.get("勤労学生") or "").strip() in ("○", "有"):
        total += 270_000

    return total


# ------------------------------------------------------------------- init


def build_input_book(path, year):
    wb = Workbook()

    ws = wb.active
    ws.title = "支払者"
    ws["A1"] = f"令和{year - 2018}年分（{year}年分）給与支払報告書  支払者の情報"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "この 1 枚は毎年ほぼ同じ。年が変わったら上の年だけ直せばよい。"
    ws["A2"].font = NOTE_FONT
    ws["A3"] = "対象年"
    ws["A3"].font = LABEL_FONT
    ws["A3"].fill = LABEL_FILL
    ws["A3"].border = BOX
    ws["B3"] = year
    ws["B3"].border = BOX

    for i, (name, note) in enumerate(PAYER_FIELDS, start=4):
        ws.cell(row=i, column=1, value=name).font = LABEL_FONT
        ws.cell(row=i, column=1).fill = LABEL_FILL
        ws.cell(row=i, column=1).border = BOX
        ws.cell(row=i, column=2).border = BOX
        if note:
            c = ws.cell(row=i, column=3, value=note)
            c.font = NOTE_FONT
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 60

    ws2 = wb.create_sheet("受給者")
    ws2["A1"] = "1 行 = 1 人。空欄でよい欄は空欄のままにする。"
    ws2["A1"].font = NOTE_FONT
    ws2["A2"] = (
        "「給与所得控除後の金額」「所得控除の額の合計額」は空欄なら自動計算する。"
        "年末調整の結果があるならそれを直接入れたほうが確実。"
    )
    ws2["A2"].font = NOTE_FONT
    ws2["A3"] = "「徴収区分」は " + " / ".join(COLLECTION_KINDS) + " のいずれか。"
    ws2["A3"].font = NOTE_FONT

    header_row = 5
    for i, (name, width, _kind) in enumerate(EMPLOYEE_FIELDS, start=1):
        c = ws2.cell(row=header_row, column=i, value=name)
        c.font = Font(size=9, bold=True)
        c.fill = LABEL_FILL
        c.alignment = CENTER
        c.border = BOX
        ws2.column_dimensions[get_column_letter(i)].width = width
    ws2.row_dimensions[header_row].height = 34

    for i, (name, _w, _k) in enumerate(EMPLOYEE_FIELDS, start=1):
        c = ws2.cell(row=header_row + 1, column=i, value=SAMPLE_EMPLOYEE.get(name))
        c.font = Font(size=10, color="9A6A00")
        c.border = BOX
    ws2.cell(row=header_row + 2, column=1, value="↑ 1 行目は記入例。実データを入れるときは上書きするか行ごと消す。")
    ws2.cell(row=header_row + 2, column=1).font = NOTE_FONT

    ws2.freeze_panes = ws2.cell(row=header_row + 1, column=3)
    wb.save(path)


# ------------------------------------------------------------------ 読み込み


def read_input_book(path):
    wb = load_workbook(path, data_only=True)

    if "支払者" not in wb.sheetnames or "受給者" not in wb.sheetnames:
        sys.exit("入力ブックに「支払者」「受給者」シートが必要です。init で作り直してください。")

    ws = wb["支払者"]
    payer = {}
    year = None
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        key, value = (row + (None, None))[:2]
        if key == "対象年":
            year = int(value) if value else None
        elif key:
            payer[str(key).strip()] = value
    if not year:
        sys.exit("支払者シートの「対象年」が空です。")

    ws2 = wb["受給者"]
    header_row = None
    for r in range(1, 20):
        if ws2.cell(row=r, column=1).value == EMPLOYEE_FIELDS[0][0]:
            header_row = r
            break
    if header_row is None:
        sys.exit("受給者シートの見出し行（受給者番号…）が見つかりません。")

    names = [ws2.cell(row=header_row, column=i).value for i in range(1, len(EMPLOYEE_FIELDS) + 1)]
    employees = []
    for r in range(header_row + 1, ws2.max_row + 1):
        values = {names[i - 1]: ws2.cell(row=r, column=i).value for i in range(1, len(names) + 1)}
        if not values.get("氏名"):
            continue
        employees.append(values)

    return payer, year, employees


def normalize(emp, year):
    """空欄を埋めて、日付を文字列にそろえる。"""
    out = dict(emp)
    for name, _w, kind in EMPLOYEE_FIELDS:
        v = out.get(name)
        if kind == "date" and v is not None and hasattr(v, "strftime"):
            out[name] = v.strftime("%Y-%m-%d")
        elif kind == "num" and isinstance(v, str):
            s = unicodedata.normalize("NFKC", v).replace(",", "").strip()
            out[name] = int(s) if s.lstrip("-").isdigit() else None

    if not out.get("種別"):
        out["種別"] = "給料・賞与"
    if not out.get("徴収区分"):
        out["徴収区分"] = TOKUBETSU
    if out.get("徴収区分") not in COLLECTION_KINDS:
        sys.exit(f"{out.get('氏名')}: 徴収区分「{out.get('徴収区分')}」は " + " / ".join(COLLECTION_KINDS) + " のいずれかにしてください。")

    if out.get("給与所得控除後の金額") in (None, ""):
        out["給与所得控除後の金額"] = salary_income(out.get("支払金額"), year)
        out["_給与所得は自動計算"] = True
    if out.get("所得控除の額の合計額") in (None, ""):
        out["所得控除の額の合計額"] = deduction_total(out)
        out["_所得控除は自動計算"] = True
    return out


# --------------------------------------------------------- 個人別明細書の描画


def draw_detail(ws, top, emp, payer, year):
    """個人別明細書を 1 枚ぶん描いて、次の開始行を返す。"""
    r = top

    cell(ws, r, 1, 16, f"令和{year - 2018}年分  給与支払報告書（個人別明細書）", align=CENTER).font = TITLE_FONT
    cell(ws, r, 17, 4, "提出先市区町村", label=True)
    cell(ws, r, 21, 4, emp.get("提出先市区町村"))
    ws.row_dimensions[r].height = 24
    r += 1

    cell(ws, r, 1, 16, "", align=CENTER)
    cell(ws, r, 17, 4, "指定番号", label=True)
    cell(ws, r, 21, 4, emp.get("指定番号") or payer.get("指定番号"))
    r += 1

    cell(ws, r, 1, 4, "支払を受ける者\n住所又は居所\n（1月1日現在）", label=True)
    cell(ws, r, 5, 3, emp.get("郵便番号"))
    cell(ws, r, 8, 12, emp.get("住所又は居所（1月1日現在）"))
    cell(ws, r, 20, 2, "徴収区分", label=True)
    cell(ws, r, 22, 3, emp.get("徴収区分"))
    ws.row_dimensions[r].height = 34
    r += 1

    cell(ws, r, 1, 3, "受給者番号", label=True)
    cell(ws, r, 4, 3, emp.get("受給者番号"))
    cell(ws, r, 7, 3, "個人番号", label=True)
    cell(ws, r, 10, 4, emp.get("個人番号"))
    cell(ws, r, 14, 2, "役職名", label=True)
    cell(ws, r, 16, 3, emp.get("役職名"))
    cell(ws, r, 19, 2, "生年月日", label=True)
    cell(ws, r, 21, 4, emp.get("生年月日"))
    r += 1

    cell(ws, r, 1, 3, "フリガナ", label=True)
    cell(ws, r, 4, 8, emp.get("フリガナ"))
    cell(ws, r, 12, 2, "氏名", label=True)
    cell(ws, r, 14, 11, emp.get("氏名"))
    r += 1

    row_of(ws, r, [(5, "種別"), (5, "支払金額"), (5, "給与所得控除後の金額"),
                   (5, "所得控除の額の合計額"), (4, "源泉徴収税額")], label=True)
    r += 1
    cell(ws, r, 1, 5, emp.get("種別"), align=CENTER)
    cell(ws, r, 6, 5, _yen(emp.get("支払金額")), money=True)
    cell(ws, r, 11, 5, _yen(emp.get("給与所得控除後の金額")), money=True)
    cell(ws, r, 16, 5, _yen(emp.get("所得控除の額の合計額")), money=True)
    cell(ws, r, 21, 4, _yen(emp.get("源泉徴収税額")), money=True)
    r += 1

    row_of(ws, r, [(3, "控除対象配偶者\nの有無等"), (3, "配偶者（特別）\n控除の額"),
                   (2, "扶養\n特定"), (2, "扶養\n老人"), (3, "うち同居老親等"),
                   (2, "扶養\nその他"), (3, "16歳未満\n扶養親族の数"),
                   (3, "特定親族\n特別控除の額"), (3, "非居住者である\n親族の数")], label=True)
    ws.row_dimensions[r].height = 28
    r += 1
    cell(ws, r, 1, 3, emp.get("控除対象配偶者の有無等"), align=CENTER)
    cell(ws, r, 4, 3, _yen(emp.get("配偶者（特別）控除の額")), money=True)
    cell(ws, r, 7, 2, _yen(emp.get("扶養親族_特定")), align=CENTER)
    cell(ws, r, 9, 2, _yen(emp.get("扶養親族_老人")), align=CENTER)
    cell(ws, r, 11, 3, _yen(emp.get("扶養親族_老人のうち同居老親等")), align=CENTER)
    cell(ws, r, 14, 2, _yen(emp.get("扶養親族_その他")), align=CENTER)
    cell(ws, r, 16, 3, _yen(emp.get("16歳未満扶養親族の数")), align=CENTER)
    cell(ws, r, 19, 3, _yen(emp.get("特定親族特別控除の額")), money=True)
    cell(ws, r, 22, 3, _yen(emp.get("非居住者である親族の数")), align=CENTER)
    r += 1

    row_of(ws, r, [(3, "障害者\n特別"), (3, "うち同居特別"), (3, "障害者\nその他"),
                   (4, "社会保険料等の金額"), (4, "うち小規模企業共済等"),
                   (4, "生命保険料の控除額"), (3, "地震保険料の控除額")], label=True)
    ws.row_dimensions[r].height = 28
    r += 1
    cell(ws, r, 1, 3, _yen(emp.get("障害者_特別（本人を除く）")), align=CENTER)
    cell(ws, r, 4, 3, _yen(emp.get("障害者_同居特別（本人を除く）")), align=CENTER)
    cell(ws, r, 7, 3, _yen(emp.get("障害者_その他（本人を除く）")), align=CENTER)
    cell(ws, r, 10, 4, _yen(emp.get("社会保険料等の金額")), money=True)
    cell(ws, r, 14, 4, _yen(emp.get("うち小規模企業共済等掛金")), money=True)
    cell(ws, r, 18, 4, _yen(emp.get("生命保険料の控除額")), money=True)
    cell(ws, r, 22, 3, _yen(emp.get("地震保険料の控除額")), money=True)
    r += 1

    cell(ws, r, 1, 4, "摘要", label=True)
    cell(ws, r, 5, 20, emp.get("摘要"))
    ws.row_dimensions[r].height = 30
    r += 1

    row_of(ws, r, [(4, "新生命保険料"), (4, "旧生命保険料"), (4, "介護医療保険料"),
                   (4, "新個人年金保険料"), (4, "旧個人年金保険料"), (4, "旧長期損害保険料")], label=True)
    r += 1
    cell(ws, r, 1, 4, _yen(emp.get("新生命保険料の金額")), money=True)
    cell(ws, r, 5, 4, _yen(emp.get("旧生命保険料の金額")), money=True)
    cell(ws, r, 9, 4, _yen(emp.get("介護医療保険料の金額")), money=True)
    cell(ws, r, 13, 4, _yen(emp.get("新個人年金保険料の金額")), money=True)
    cell(ws, r, 17, 4, _yen(emp.get("旧個人年金保険料の金額")), money=True)
    cell(ws, r, 21, 4, _yen(emp.get("旧長期損害保険料の金額")), money=True)
    r += 1

    row_of(ws, r, [(3, "住宅借入金等\n特別控除の額"), (3, "適用数"), (4, "居住開始年月日1"),
                   (3, "区分1"), (4, "年末残高1"), (4, "居住開始年月日2"), (3, "年末残高2")], label=True)
    ws.row_dimensions[r].height = 26
    r += 1
    cell(ws, r, 1, 3, _yen(emp.get("住宅借入金等特別控除の額")), money=True)
    cell(ws, r, 4, 3, _yen(emp.get("住宅借入金等特別控除適用数")), align=CENTER)
    cell(ws, r, 7, 4, emp.get("居住開始年月日1"), align=CENTER)
    cell(ws, r, 11, 3, emp.get("住宅借入金等特別控除区分1"), align=CENTER)
    cell(ws, r, 14, 4, _yen(emp.get("住宅借入金等年末残高1")), money=True)
    cell(ws, r, 18, 4, emp.get("居住開始年月日2"), align=CENTER)
    cell(ws, r, 22, 3, _yen(emp.get("住宅借入金等年末残高2")), money=True)
    r += 1

    row_of(ws, r, [(4, "配偶者の合計所得"), (4, "国民年金保険料等の金額"),
                   (3, "基礎控除の額"), (4, "所得金額調整控除額"),
                   (3, "本人が障害者"), (2, "寡婦"), (2, "ひとり親"), (2, "勤労学生")], label=True)
    ws.row_dimensions[r].height = 26
    r += 1
    cell(ws, r, 1, 4, _yen(emp.get("配偶者の合計所得")), money=True)
    cell(ws, r, 5, 4, _yen(emp.get("国民年金保険料等の金額")), money=True)
    cell(ws, r, 9, 3, _yen(emp.get("基礎控除の額")), money=True)
    cell(ws, r, 12, 4, _yen(emp.get("所得金額調整控除額")), money=True)
    cell(ws, r, 16, 3, emp.get("本人が障害者"), align=CENTER)
    cell(ws, r, 19, 2, emp.get("寡婦"), align=CENTER)
    cell(ws, r, 21, 2, emp.get("ひとり親"), align=CENTER)
    cell(ws, r, 23, 2, emp.get("勤労学生"), align=CENTER)
    r += 1

    row_of(ws, r, [(2, "未成年者"), (2, "外国人"), (2, "死亡退職"), (2, "災害者"), (2, "乙欄"),
                   (4, "中途就職年月日"), (4, "中途退職年月日"), (6, "受給者生年月日")], label=True)
    r += 1
    cell(ws, r, 1, 2, emp.get("未成年者"), align=CENTER)
    cell(ws, r, 3, 2, emp.get("外国人"), align=CENTER)
    cell(ws, r, 5, 2, emp.get("死亡退職"), align=CENTER)
    cell(ws, r, 7, 2, emp.get("災害者"), align=CENTER)
    cell(ws, r, 9, 2, emp.get("乙欄"), align=CENTER)
    cell(ws, r, 11, 4, emp.get("中途就職年月日"), align=CENTER)
    cell(ws, r, 15, 4, emp.get("中途退職年月日"), align=CENTER)
    cell(ws, r, 19, 6, emp.get("生年月日"), align=CENTER)
    r += 1

    cell(ws, r, 1, 4, "支払者\n個人番号又は法人番号", label=True)
    cell(ws, r, 5, 6, payer.get("個人番号又は法人番号"))
    cell(ws, r, 11, 3, "所在地", label=True)
    cell(ws, r, 14, 11, payer.get("所在地"))
    ws.row_dimensions[r].height = 26
    r += 1

    cell(ws, r, 1, 4, "名称", label=True)
    cell(ws, r, 5, 8, payer.get("名称"))
    cell(ws, r, 13, 3, "代表者職氏名", label=True)
    cell(ws, r, 16, 5, payer.get("代表者職氏名"))
    cell(ws, r, 21, 2, "電話", label=True)
    cell(ws, r, 23, 2, payer.get("電話番号"))
    r += 1

    notes = []
    if emp.get("_給与所得は自動計算"):
        notes.append("給与所得控除後の金額は支払金額から自動計算した参考値")
    if emp.get("_所得控除は自動計算"):
        notes.append("所得控除の額の合計額は入力された控除額の合計（参考値）")
    if notes:
        c = cell(ws, r, 1, FORM_COLS, "※ " + " / ".join(notes) + "。年末調整の結果と必ず突き合わせること。")
        c.font = NOTE_FONT
        r += 1

    return r + 2  # 次の明細まで 2 行あける


# --------------------------------------------------------------- 総括表の描画


def draw_summary(ws, top, city, emps, payer, year):
    r = top

    cell(ws, r, 1, 18, f"令和{year - 2018}年分  給与支払報告書（総括表）", align=CENTER).font = TITLE_FONT
    ws.row_dimensions[r].height = 26
    r += 1

    cell(ws, r, 1, 4, "提出先", label=True)
    cell(ws, r, 5, 8, f"{city} 長 殿")
    cell(ws, r, 13, 2, "指定番号", label=True)
    indicated = next((e.get("指定番号") for e in emps if e.get("指定番号")), payer.get("指定番号"))
    cell(ws, r, 15, 4, indicated)
    r += 1

    for label, value, span in [
        ("個人番号又は法人番号", payer.get("個人番号又は法人番号"), 14),
        ("所在地", f"〒{payer.get('所在地郵便番号') or ''}  {payer.get('所在地') or ''}".strip(), 14),
        ("名称", payer.get("名称"), 14),
        ("代表者職氏名", payer.get("代表者職氏名"), 14),
        ("電話番号", payer.get("電話番号"), 14),
        ("事業種目", payer.get("事業種目"), 14),
        ("所轄税務署", payer.get("所轄税務署"), 14),
        ("給与支払の方法及び期日", payer.get("給与支払の方法及び期日"), 14),
        ("納入書の送付", payer.get("納入書の送付"), 14),
    ]:
        cell(ws, r, 1, 4, label, label=True)
        cell(ws, r, 5, span, value)
        r += 1

    contact = "  ".join(
        str(v) for v in [payer.get("連絡者の係名"), payer.get("連絡者の氏名"), payer.get("連絡者の電話番号")] if v
    )
    cell(ws, r, 1, 4, "連絡者", label=True)
    cell(ws, r, 5, 14, contact)
    r += 1

    zeirishi = "  ".join(
        str(v) for v in [payer.get("関与税理士等の氏名"), payer.get("関与税理士等の電話番号")] if v
    )
    cell(ws, r, 1, 4, "関与税理士等", label=True)
    cell(ws, r, 5, 14, zeirishi)
    r += 2

    counts = {k: sum(1 for e in emps if e.get("徴収区分") == k) for k in COLLECTION_KINDS}
    cell(ws, r, 1, 4, "受給者総人員", label=True)
    cell(ws, r, 5, 3, len(emps), align=CENTER)
    cell(ws, r, 8, 11, "（1月1日現在。退職者を含む在職者数）")
    ws.cell(row=r, column=8).font = NOTE_FONT
    r += 1

    row_of(ws, r, [(4, "報告人員"), (4, TOKUBETSU), (5, FUTSU_TAISHOKU),
                   (5, FUTSU_SONOTA), (2, "計")], label=True)
    ws.row_dimensions[r].height = 30
    r += 1
    cell(ws, r, 1, 4, "人数", label=True)
    cell(ws, r, 5, 4, counts[TOKUBETSU], align=CENTER)
    cell(ws, r, 9, 5, counts[FUTSU_TAISHOKU], align=CENTER)
    cell(ws, r, 14, 5, counts[FUTSU_SONOTA], align=CENTER)
    cell(ws, r, 19, 2, len(emps), align=CENTER)
    r += 2

    cell(ws, r, 1, 18, "内訳（この総括表に添付する個人別明細書）", label=True, align=LEFT)
    r += 1
    row_of(ws, r, [(3, "受給者番号"), (5, "氏名"), (5, "徴収区分"),
                   (5, "支払金額")], label=True)
    r += 1
    for e in sorted(emps, key=lambda x: str(x.get("受給者番号") or "")):
        cell(ws, r, 1, 3, e.get("受給者番号"), align=CENTER)
        cell(ws, r, 4, 5, e.get("氏名"))
        cell(ws, r, 9, 5, e.get("徴収区分"), align=CENTER)
        cell(ws, r, 14, 5, _yen(e.get("支払金額")), money=True)
        r += 1
    cell(ws, r, 1, 13, "合計", label=True)
    cell(ws, r, 14, 5, sum(int(e.get("支払金額") or 0) for e in emps), money=True)
    r += 1

    return r + 2


# ------------------------------------------------------------------- build


def build_report(in_path, out_path):
    payer, year, raw = read_input_book(in_path)
    if not raw:
        sys.exit("受給者シートにデータがありません。")
    employees = [normalize(e, year) for e in raw]

    wb = Workbook()

    ws = wb.active
    ws.title = "個人別明細書"
    for i in range(1, FORM_COLS + 1):
        ws.column_dimensions[get_column_letter(i)].width = 5.2
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    r = 1
    for i, emp in enumerate(employees):
        if i:
            ws.row_breaks.append(Break(id=r - 1))
        r = draw_detail(ws, r, emp, payer, year)

    ws2 = wb.create_sheet("総括表")
    for i in range(1, 19):
        ws2.column_dimensions[get_column_letter(i)].width = 7
    ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
    ws2.page_setup.fitToWidth = 1
    ws2.page_setup.fitToHeight = 0
    ws2.sheet_properties.pageSetUpPr.fitToPage = True

    cities = {}
    for e in employees:
        cities.setdefault(e.get("提出先市区町村") or "（市区町村未入力）", []).append(e)

    r = 1
    for i, (city, emps) in enumerate(sorted(cities.items())):
        if i:
            ws2.row_breaks.append(Break(id=r - 1))
        r = draw_summary(ws2, r, city, emps, payer, year)

    wb.save(out_path)
    return year, employees, cities


def main():
    p = argparse.ArgumentParser(description="給与支払報告書を Excel で作る")
    sub = p.add_subparsers(dest="cmd", required=True)

    p_init = sub.add_parser("init", help="入力用ブックを作る")
    p_init.add_argument("out")
    p_init.add_argument("--year", type=int, default=2025, help="対象年（西暦）")

    p_build = sub.add_parser("build", help="入力用ブックから報告書を作る")
    p_build.add_argument("src")
    p_build.add_argument("--out", required=True)

    args = p.parse_args()

    if args.cmd == "init":
        build_input_book(args.out, args.year)
        print(f"入力用ブックを作りました: {args.out}（対象年 {args.year}）")
        return

    year, employees, cities = build_report(args.src, Path(args.out))
    print(f"報告書を作りました: {args.out}")
    print(f"  対象年 令和{year - 2018}年分（{year}年分）")
    print(f"  個人別明細書 {len(employees)} 枚 / 総括表 {len(cities)} 枚（{'、'.join(sorted(cities))}）")
    for kind in COLLECTION_KINDS:
        n = sum(1 for e in employees if e.get("徴収区分") == kind)
        if n:
            print(f"  {kind}: {n} 人")
    auto = [e["氏名"] for e in employees if e.get("_給与所得は自動計算") or e.get("_所得控除は自動計算")]
    if auto:
        print("  ※ 自動計算で埋めた欄がある受給者: " + "、".join(auto))


if __name__ == "__main__":
    main()
